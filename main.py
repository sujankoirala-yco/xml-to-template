
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font, Border, Side
import os
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv()

# Access variables
user = os.getenv("USER")
password = os.getenv("PASSWORD")
host = os.getenv("HOST")
port = os.getenv("PORT")
dbname = os.getenv("DBNAME")

def export_postgres_query_to_template():

    engine = create_engine(f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{dbname}")

    query = """
    select ORG_NUM, CHANNEL, DIVISION, 
    DLY_TY_NET_SLS_AMT, DLY_LY_NET_SLS_AMT, DLY_NET_SLS_VAR,
    DLY_TY_CHANNEL_NET_SLS, DLY_LY_CHANNEL_NET_SLS, DLY_CHANNEL_NET_SLS_VAR,
    DLY_TY_NET_PROFIT_AMT, DLY_LY_NET_PROFIT_AMT, DLY_NET_PROFIT_VAR,
    DLY_TY_CHANNEL_NET_PROFIT, DLY_LY_CHANNEL_NET_PROFIT, DLY_CHANNEL_PROFIT_VAR,
    DLY_TY_CHANNEL_BUYING, DLY_LY_CHANNEL_BUYING, DLY_CHANNEL_BUYING_VAR,
    DLY_TY_DIV_CHAN_BUYING, DLY_LY_DIV_CHAN_BUYING, DLY_DIV_CHAN_BUYING_VAR,
<<<<<<< HEAD
    DLY_TY_CHANNEL_INVOICE FROM sales_data;
=======
    DLY_TY_CHANNEL_INVOICE, DLY_LY_CHANNEL_INVOICE, DLY_CHANNEL_INVOICE_VAR,
    DLY_TY_DIV_CHAN_INVOICE, DLY_LY_DIV_CHAN_INVOICE, DLY_DIV_CHAN_INVOICE_VAR,
    DLY_LY_DIV_AVG_INVOICE, DLY_TY_DIV_AVG_INVOICE, DLY_DIV_AVG_INVOICE_VAR FROM sales_data;
>>>>>>> d1c66c8 (feat(postgre-excel): implement loop for static cell mapping)
    """

    template_path = "./new_xlsx_template.xlsx"
    output_path = "./src/public/query_result_filled.xlsx"

    # cell_map = {
    #     'money': {
    #         'start_cell': 'C10',
    #         'font_size': 12,
    #         'row_height': 20,
    #         'border': True
    #     },
    #     'age_no': {
    #         # 'start_cell': 'D10',
    #         # 'font_size': 11,
    #         # 'row_height': 18,
    #         # 'border': False
    #         'start_cell': 'D10',
    #         'font_size': 12,
    #         'row_height': 20,
    #         'border': True
    #     }
    # }


    # List of columns from your SELECT query
    columns = [

        'DLY_TY_NET_SLS_AMT', 'DLY_LY_NET_SLS_AMT', 'DLY_NET_SLS_VAR',
        'DLY_TY_CHANNEL_NET_SLS', 'DLY_LY_CHANNEL_NET_SLS', 'DLY_CHANNEL_NET_SLS_VAR',
        'DLY_TY_NET_PROFIT_AMT', 'DLY_LY_NET_PROFIT_AMT', 'DLY_NET_PROFIT_VAR',
        'DLY_TY_CHANNEL_NET_PROFIT', 'DLY_LY_CHANNEL_NET_PROFIT', 'DLY_CHANNEL_PROFIT_VAR',
        'DLY_TY_CHANNEL_BUYING', 'DLY_LY_CHANNEL_BUYING', 'DLY_CHANNEL_BUYING_VAR',
        'DLY_TY_DIV_CHAN_BUYING', 'DLY_LY_DIV_CHAN_BUYING', 'DLY_DIV_CHAN_BUYING_VAR',
<<<<<<< HEAD
        'DLY_TY_CHANNEL_INVOICE'
=======
        'DLY_TY_CHANNEL_INVOICE', 'DLY_LY_CHANNEL_INVOICE', 'DLY_CHANNEL_INVOICE_VAR',
    'DLY_TY_DIV_CHAN_INVOICE', 'DLY_LY_DIV_CHAN_INVOICE', 'DLY_DIV_CHAN_INVOICE_VAR'
>>>>>>> d1c66c8 (feat(postgre-excel): implement loop for static cell mapping)
    ]

    # Starting row
    start_row = 10

    # Start from column C (3rd column in Excel)
    start_col_index = 3

    cell_map = {}

    for i, col_name in enumerate(columns):
        col_letter = get_column_letter(start_col_index + i)
        cell_map[col_name] = {
            'start_cell': f"{col_letter}{start_row}",
            'font_size': 12,
            'row_height': 20,
            'border': True
        }

    # # Optional: print to verify
    # for k, v in cell_map.items():
    #     print(k, v)
    print(cell_map)



    conn = None
    try:
        print("🔹 Connecting to PostGreSQL...")
        df = pd.read_sql(query, engine)
        df.columns = [c.upper() for c in df.columns]

        # print(df.head())

        print("🔹 Executing query...")
        print("✅ Query executed successfully")
        print(f"🔹 Number of rows fetched: {len(df)}")
        print("🔹 Columns fetched:", df.columns.tolist())

        if not os.path.exists(template_path):
            print(f"⚠️ Template file '{template_path}' not found.")
            return

        print(f"🔹 Loading Excel template: {template_path}")
        wb = load_workbook(template_path)
        ws = wb.active
        print(f"✅ Template loaded, active sheet: {ws.title}")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_name, options in cell_map.items():
            start_cell = options.get('start_cell', 'A1')
            font_size = options.get('font_size', 11)
            row_height = options.get('row_height', 15)
            apply_border = options.get('border', False)

            print(f"🔹 Processing column '{col_name}' into cell '{start_cell}'")

            if col_name not in df.columns:
                print(f"⚠️ Column '{col_name}' not found in query results, skipping.")
                continue

            start_col = column_index_from_string(''.join(filter(str.isalpha, start_cell)))
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            print(f"🔹 Starting at row {start_row}, column {start_col}")

            for i, value in enumerate(df[col_name], start=start_row):
                cell = ws.cell(row=i, column=start_col, value=value)

                # Apply font size
                cell.font = Font(size=font_size)

                # Apply border if enabled
                if apply_border:
                    cell.border = thin_border

                # Apply row height
                ws.row_dimensions[i].height = row_height

                if i < start_row + 5:
                    print(f"  - Writing row {i}, col {start_col}: {value}")

        print(f"🔹 Saving filled template to: {output_path}")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"✅ Data written to template and saved as '{output_path}'")


    except Exception as e:
        print(f"❌ Unexpected error: {e}")

    finally:
        if conn:
            conn.close()
            print("🔹 MySQL connection closed")


if __name__ == "__main__":
    export_postgres_query_to_template()
